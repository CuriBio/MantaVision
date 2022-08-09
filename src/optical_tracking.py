import argparse
import os
import shutil
import sys
import json
import time
from typing import Tuple, List, Dict
import zipfile
import openpyxl
import cv2 as cv
from datetime import datetime
from os_functions import contentsOfDir, getFilePathViaGUI, getDirPathViaGUI
from video2images import video2images
from track_template import trackTemplate
from video_api import supported_file_extensions


# TODO: separate out the file system functions like contentsOfDir, fileSystemViaGUI etc
#       put them into a file called "os_functions.py"

# TODO: in the user xlsx, have a pixel displacement column and converted displacement column

# TODO: for contraction measurements of regular videos
#       we can use tracking to determine the smallest min and largest max positions
#       that way we'd have a good approximation of where the relaxation and contractions should be
#       and we can then pick out the time stamps of frames where there is a local min/max that is
#       within some small margin of error (say +/-10% of the difference in (largest max - smallest min)
#       using these time stamps for known min/max, we can plot the peaks and troughs and estimate
#       the frequency etc.


def runTrackTemplate(config: Dict):
    track_templates_start_time = time.time()
    dirs, args = verifiedInputs(config)

    if 'errors' in args:
        if 'input dir empty' in args['errors']:
            print('WARNING. The selected directory is empty. Nothing to do. Exiting')
            return

    # make all the dirs that are needed for writing the results and barf if any dirs already exist
    dirs_exist_error_message = ''
    if os.path.isdir(dirs['results_dir_path']):
        dirs_exist_error_message += "results dir already exists. Cannot overwrite.\n"
    if os.path.isdir(dirs['results_json_dir_path']):
        dirs_exist_error_message += "json results dir already exists. Cannot overwrite.\n"
    if os.path.isdir(dirs['results_xlsx_dir_path']):
        dirs_exist_error_message += "xlsx results dir already exists. Cannot overwrite.\n"
    if os.path.isdir(dirs['results_video_dir_path']):
        dirs_exist_error_message += "video results dir already exists. Cannot overwrite.\n"
    if os.path.isdir(dirs['results_template_dir_path']):
        dirs_exist_error_message += "template results dir already exists. Cannot overwrite.\n"
    if len(dirs_exist_error_message) > 0:
        dirs_exist_error_message = "ERROR.\n" + dirs_exist_error_message + "Nothing Tracked."
        print(dirs_exist_error_message)
        sys.exit(1)
    os.mkdir(dirs['results_dir_path'])
    os.mkdir(dirs['results_json_dir_path'])
    os.mkdir(dirs['results_xlsx_dir_path'])
    os.mkdir(dirs['results_video_dir_path'])
    os.mkdir(dirs['results_template_dir_path'])
    os.mkdir(dirs['results_video_min_frames_dir_path'])
    if next(iter(args))['output_frames']:
        os.mkdir(dirs['results_video_frames_dir_path'])

    # run the tracking routine on each input video
    # and write out the results
    print("\nTemplate Tracker running...")
    template_image = None
    total_tracking_time = 0
    for input_args in args:
        print(f'processing: {input_args["input_video_path"]}')
        video_tracking_start_time = time.time()
        messages, tracking_results, frames_per_second, template, min_frame_number = trackTemplate(
            input_args['input_video_path'],
            input_args['template_guide_image_path'],
            template_image,
            input_args['output_video_path'],
            input_args['guide_match_search_seconds'],
            input_args['microns_per_pixel'],
            input_args['output_conversion_factor'],
            input_args['sub_pixel_search_increment'],
            input_args['sub_pixel_refinement_radius'],
            input_args['user_roi_selection'],
            input_args['max_translation_per_frame'],
            input_args['max_rotation_per_frame'],
            input_args['contraction_vector']
        )
        total_tracking_time += (time.time() - video_tracking_start_time)

        # check for any errors
        warning_msg, error_msg = messages
        if warning_msg is not None:
            print(warning_msg)
        if error_msg is not None:
            print(error_msg)
            sys.exit(1)

        # write the template used for tracking to the results dir
        cv.imwrite(input_args['results_template_filename'], template)

        # write out the frame with the min movement position
        video2images(
            input_video_path=input_args['input_video_path'],
            output_dir_path=dirs['results_video_min_frames_dir_path'],
            enhance_contrast=False,
            frame_number_to_write=min_frame_number,
            image_extension='tiff'
        )

        # write out the results video as frames if requested
        if input_args['output_frames']:
            os.mkdir(input_args['output_video_frames_dir_path'])
            video2images(
                input_video_path=input_args['output_video_path'],
                output_dir_path=input_args['output_video_frames_dir_path'],
                enhance_contrast=False,
                image_extension='jpg'  # don't need high quality images for this
            )

        # write results to xlsx files
        meta_data = {
            'Well Name': input_args['well_name'],
            'Date Stamp': input_args['date_stamp'],
            'Frames Per Second': frames_per_second,
            'User ROI Selection': input_args['user_roi_selection'],
            'Max Translation Per Frame': input_args['max_translation_per_frame'],
            'Max Rotation Per Frame': input_args['max_rotation_per_frame'],
            'Contraction Vector': input_args['contraction_vector'],
            'Microns Per Pixel': input_args['microns_per_pixel'],
            'Output Conversion Factor': input_args['output_conversion_factor'],
            'Sub Pixel Search Increment': input_args['sub_pixel_search_increment'],
            'Sub Pixel Refinement Radius': input_args['sub_pixel_refinement_radius']
        }
        # fill in missing values in meta_data
        for md_key, md_value in meta_data.items():
            if md_value is None:
                meta_data[md_key] = 'None'
        resultsToCSVforUser(
            tracking_results,
            meta_data,
            input_args['path_to_user_excel_results']
        )
        resultsToCSVforSDK(
            tracking_results,
            meta_data,
            input_args['path_to_excel_template'],
            input_args['path_to_sdk_excel_results']
        )

        # write the run config and results as json
        if input_args['output_json_path'] is not None:
            tracking_results_complete = {
                "INPUT_ARGS": input_args,
                "ERROR_MSGS": error_msg,
                "WARNING_MSGS": warning_msg,
                "RESULTS": tracking_results,
            }
            with open(input_args['output_json_path'], 'w') as outfile:
                json.dump(tracking_results_complete, outfile, indent=4)

    # create a zip archive and write all the xlsx files to it
    xlsx_archive_file_path = os.path.join(dirs['results_dir_path'], 'xlsx-results.zip')
    xlsx_archive = zipfile.ZipFile(xlsx_archive_file_path, 'w')
    for dir_name, _, file_names in os.walk(dirs['results_xlsx_dir_path']):
        for file_name in file_names:
            if 'sdk' in file_name:  # only include the xlsx files intended for the sdk
                file_path = os.path.join(dir_name, file_name)
                xlsx_archive.write(file_path, os.path.basename(file_path))
    xlsx_archive.close()

    num_videos_processed = len(args)
    track_templates_runtime = time.time() - track_templates_start_time
    per_video_tracking_time: float = float(total_tracking_time) / float(num_videos_processed)
    print(f'...Template Tracker completed in {round(track_templates_runtime, 2)}s')
    total_tracking_time = round(total_tracking_time, 2)
    per_video_tracking_time = round(per_video_tracking_time, 2)
    print(f'\nActual tracking time for {num_videos_processed} videos: {total_tracking_time}s')
    print(f'{per_video_tracking_time}s per video')


def verifiedInputs(config: Dict) -> Tuple[str, List[Dict]]:
    """ """
    error_msgs = []
    # check the dir path to input videos
    open_video_dir_dialog = False
    if 'input_video_path' in config:
        if config['input_video_path'] is None:
            open_video_dir_dialog = True
        else:
            if config['input_video_path'].lower() == 'select_dir':
                open_video_dir_dialog = True
            elif not os.path.isdir(config['input_video_path']):
                error_msgs.append('Input path to video/s does not exist.')
    else:
        open_video_dir_dialog = True
    # pop up a dialog to select the dir for videos if required
    if open_video_dir_dialog:
        print()
        print("waiting for user input (video dir) via pop up dialog box...")
        dir_path_via_gui = getDirPathViaGUI(window_title='Select Directory With Videos To Track')
        if dir_path_via_gui == () or dir_path_via_gui == '':
            error_msgs.append('No input path to video/s was provided.')
        else:
            config['input_video_path'] = dir_path_via_gui
        print("...user input obtained from pop up dialog box.")

    # check the file path to the template image
    user_roi_selection = False
    open_template_dir_dialog = False
    if 'template_guide_image_path' in config:
        if config['template_guide_image_path'] is None:
            config['template_guide_image_path'] = ''
            user_roi_selection = True
        else:
            if config['template_guide_image_path'].lower() == 'select_file':
                open_template_dir_dialog = True
            elif config['template_guide_image_path'].lower() == 'draw':
                config['template_guide_image_path'] = ''
                user_roi_selection = True
            elif config['template_guide_image_path'].lower() == '':
                user_roi_selection = True
            elif not os.path.isfile(config['template_guide_image_path']):
                error_msgs.append('Input template image file does not exist.')
    else:
        user_roi_selection = True
    # pop up a dialog to select the template file if required
    if open_template_dir_dialog:
        print()
        print("waiting for user input (template to use) via pop up dialog box...")
        file_path_via_gui = getFilePathViaGUI(window_title='Select File With Template To Track')
        if file_path_via_gui == () or file_path_via_gui == '':
            error_msgs.append('No input template image path was provided.')
        else:
            config['template_guide_image_path'] = file_path_via_gui
        print("...user input obtained from pop up dialog box.")

    # barf if there was an error with either the input video dir path or template file path
    if len(error_msgs) > 0:
        error_msg = 'ERROR.'
        for error_string in error_msgs:
            error_msg = error_msg + ' ' + error_string
        error_msg += ' Nothing to do. Exiting.'
        print(error_msg)
        sys.exit(1)

    template_guide_image_path = config['template_guide_image_path']
    base_dir, video_files = contentsOfDir(dir_path=config['input_video_path'], search_terms=supported_file_extensions)
    if video_files is None:
        return None, {'errors': {'input dir empty'}}
    else:
        num_videos_in_dir = len(video_files)
        if num_videos_in_dir < 1:
            return None, {'errors': {'input dir empty'}}

    unique_name = "results_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    results_dir_path = os.path.join(base_dir, unique_name)
    results_json_dir_path = os.path.join(results_dir_path, 'json')
    results_xlsx_dir_path = os.path.join(results_dir_path, 'xlsx')
    results_video_dir_path = os.path.join(results_dir_path, 'video')
    results_template_dir_path = os.path.join(results_dir_path, 'template')
    results_video_frames_dir_path = os.path.join(results_video_dir_path, 'frames')
    results_video_min_frames_dir_path = os.path.join(results_dir_path, 'min_frame')

    dirs = {
        'base_dir': base_dir,
        'results_dir_path': results_dir_path,
        'results_json_dir_path': results_json_dir_path,
        'results_xlsx_dir_path': results_xlsx_dir_path,
        'results_video_dir_path': results_video_dir_path,
        'results_template_dir_path': results_template_dir_path,
        'results_video_frames_dir_path': results_video_frames_dir_path,
        'results_video_min_frames_dir_path': results_video_min_frames_dir_path
    }

    if 'path_to_excel_template' not in config:
        path_to_excel_template = None
    else:
        path_to_excel_template = config['path_to_excel_template']

    if 'guide_match_search_seconds' not in config:
        guide_match_search_seconds = None
    else:
        guide_match_search_seconds = config['guide_match_search_seconds']

    if 'microns_per_pixel' not in config:
        microns_per_pixel = None
    else:
        microns_per_pixel = config['microns_per_pixel']

    if 'output_conversion_factor' not in config:
        output_conversion_factor = None
    else:
        output_conversion_factor = config['output_conversion_factor']

    if 'sub_pixel_search_increment' not in config:
        sub_pixel_search_increment = None
    else:
        sub_pixel_search_increment = config['sub_pixel_search_increment']

    if 'sub_pixel_refinement_radius' not in config:
        sub_pixel_refinement_radius = None
    else:
        sub_pixel_refinement_radius = config['sub_pixel_refinement_radius']
    if sub_pixel_search_increment is None and sub_pixel_refinement_radius is not None:
        print('WARNING. sub_pixel_refinement_radius ignored because sub_pixel_search_increment not provided')
        sub_pixel_refinement_radius = None

    if 'max_translation_per_frame' not in config:
        max_translation_per_frame = None
    elif config['max_translation_per_frame'] is None:
        max_translation_per_frame = None
    else:
        max_translation_per_frame = (config['max_translation_per_frame'], config['max_translation_per_frame'])

    if 'max_rotation_per_frame' not in config:
        max_rotation_per_frame = None
    else:
        max_rotation_per_frame = config['max_rotation_per_frame']

    if 'output_frames' not in config:
        output_frames = False
    else:
        output_frames = config['output_frames']

    if 'contraction_vector' not in config:
        contraction_vector = None
    else:
        contraction_vector = config['contraction_vector']

    # set all the values needed to run template matching on each input video
    configs = []
    for file_name, input_file_extension in video_files:
        # if the file is long enough to contain well name and data stamp,
        # extract them from expect
        well_name = 'A001'
        date_stamp = '2020-02-02'
        num_chars_in_datestamp = len(date_stamp)
        num_chars_in_well_name = len(well_name)
        min_num_chars_in_file_name = num_chars_in_datestamp + num_chars_in_well_name
        if len(file_name) >= min_num_chars_in_file_name:
            file_name_length = len(file_name)
            well_name = file_name[file_name_length - num_chars_in_well_name:]
            date_stamp = file_name[:num_chars_in_datestamp]

        # set all the required path values
        input_video_path = os.path.join(base_dir, file_name + input_file_extension)
        output_video_frames_dir_path = os.path.join(results_video_frames_dir_path, file_name)
        output_json_path = os.path.join(results_json_dir_path, file_name + '-results.json')
        path_to_user_excel_results = os.path.join(results_xlsx_dir_path, file_name + '-reslts_user.xlsx')
        path_to_sdk_excel_results = os.path.join(results_xlsx_dir_path, file_name + '-reslts_sdk.xlsx')
        output_file_extension = '.mp4'
        output_video_path = os.path.join(results_video_dir_path, file_name + '-results' + output_file_extension)
        results_template_filename = os.path.join(results_template_dir_path, file_name + '-template.tiff')

        configs.append({
            'input_video_path': input_video_path,
            'template_guide_image_path': template_guide_image_path,
            'results_template_filename': results_template_filename,
            'user_roi_selection': user_roi_selection,
            'output_video_path': output_video_path,
            'output_video_frames_dir_path': output_video_frames_dir_path,
            'output_json_path': output_json_path,
            'path_to_excel_template': path_to_excel_template,
            'path_to_sdk_excel_results': path_to_sdk_excel_results,
            'path_to_user_excel_results': path_to_user_excel_results,
            'guide_match_search_seconds': guide_match_search_seconds,
            'microns_per_pixel': microns_per_pixel,
            'output_conversion_factor': output_conversion_factor,
            'sub_pixel_search_increment': sub_pixel_search_increment,
            'sub_pixel_refinement_radius': sub_pixel_refinement_radius,
            'max_translation_per_frame': max_translation_per_frame,
            'max_rotation_per_frame': max_rotation_per_frame,
            'output_frames': output_frames,
            'contraction_vector': contraction_vector,
            'well_name': well_name,
            'date_stamp': date_stamp,
        })

    return (dirs, configs)


def resultsToCSVforSDK(
        tracking_results: List[Dict],
        meta_data: Dict,
        path_to_template_file: str,
        path_to_output_file,
):
    # create a file from scratch or template
    if path_to_template_file is None:
        workbook = openpyxl.Workbook()  # open a blank workbook
    else:  # open the template workbook
        shutil.copyfile(path_to_template_file, path_to_output_file)
        workbook = openpyxl.load_workbook(filename=path_to_output_file)
    sheet = workbook.active

    # add meta data
    well_name = meta_data['Well Name']
    if well_name is None:
        well_name = 'Z01'
    sheet['E2'] = well_name
    date_stamp = meta_data['Date Stamp']
    sheet['E3'] = date_stamp + ' 00:00:00'
    sheet['E4'] = 'NA'  # plate barcode
    frames_per_second = meta_data['Frames Per Second']
    sheet['E5'] = frames_per_second
    sheet['E6'] = 'y'  # do twiches point up
    sheet['E7'] = 'NA'  # microscope name

    # add runtime data (time, displacement etc)
    template_start_row = 2
    time_column = 'A'
    displacement_column = 'B'
    num_rows_to_write = len(tracking_results)
    for results_row in range(num_rows_to_write):
        tracking_result = tracking_results[results_row]
        sheet_row = str(results_row + template_start_row)
        sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
        sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
    workbook.save(filename=path_to_output_file)


def resultsToCSVforUser(tracking_results: List[Dict], meta_data: Dict, path_to_output_file: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # add meta data
    meta_data_name_column = 'G'
    meta_data_value_column = 'H'
    meta_data_row_number = 1
    sheet[meta_data_value_column + str(meta_data_row_number)] = 'Runtime Meta Data'
    for column_name, column_value in meta_data.items():
        meta_data_row_number += 1
        sheet[meta_data_name_column + str(meta_data_row_number)] = column_name
        if isinstance(column_value, (int, float)):
            sheet[meta_data_value_column + str(meta_data_row_number)] = column_value
        else:
            sheet[meta_data_value_column + str(meta_data_row_number)] = f"{column_value}"

    # add runtime data
    # heading fields
    heading_row = 1
    time_column = 'A'
    sheet[time_column + str(heading_row)] = 'Time (s)'
    displacement_column = 'B'
    sheet[displacement_column + str(heading_row)] = 'Displacement From Min'
    x_pos_column = 'C'
    sheet[x_pos_column + str(heading_row)] = 'Template Match Center X (pixel pos)'
    y_pos_column = 'D'
    sheet[y_pos_column + str(heading_row)] = 'Template Match Center Y (pixel pos)'
    angle_column = 'E'
    sheet[angle_column + str(heading_row)] = 'Template Match Angle (deg)'

    # time, displacement from ref position, absolute position and angle fields
    data_row = heading_row + 1
    num_rows_to_write = len(tracking_results)
    for results_row in range(num_rows_to_write):
        tracking_result = tracking_results[results_row]
        sheet_row = str(results_row + data_row)
        sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
        sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
        sheet[x_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_X'])
        sheet[y_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_Y'])
        sheet[angle_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ROTATION'])
    workbook.save(filename=path_to_output_file)


def config_from_json(json_config_path) -> Tuple[str, List[Dict]]:
    json_file = open(json_config_path)
    config = json.load(json_file)
    return config


def config_from_cmdline(cmdline_args) -> dict:
    return {
        'input_video_path': cmdline_args.input_video_path,
        'template_guide_image_path': cmdline_args.template_guide_image_path,
        'output_path': cmdline_args.output_path,
        'guide_match_search_seconds': cmdline_args.guide_match_search_seconds,
        'microns_per_pixel': cmdline_args.microns_per_pixel,
        'path_to_excel_template': cmdline_args.path_to_excel_template
    }


def main():
    # read in a config file & parse the input args
    parser = argparse.ArgumentParser(
        description='Tracks a template image through each frame of a video.',
    )
    parser.add_argument(
        '--input_video_path',
        default=None,
        help='Path of input video/s to track a template.',
    )
    parser.add_argument(
        '--template_guide_image_path',
        default=None,
        help='Path to an image that will be used as a template to match.',
    )
    parser.add_argument(
        '--path_to_excel_template',
        default=None,
        help='path to exel spread sheet used as a template to write the results into',
    )
    parser.add_argument(
        '--json_config_path',
        default=None,
        help='Path of a json file with run config parameters'
    )
    parser.add_argument(
        '--output_path',
        default=None,
        help='Path to write tracking results.',
    )
    parser.add_argument(
        '--guide_match_search_seconds',
        default=None,
        help='number of seconds to search the video for the best match with the guide template.',
    )
    parser.add_argument(
        '--microns_per_pixel',
        default=None,
        help='conversion factor for pixel distances to microns',
    )
    raw_args = parser.parse_args()
    if raw_args.json_config_path is not None:
        config = config_from_json(raw_args.json_config_path)
    else:
        config = config_from_cmdline(raw_args)
    runTrackTemplate(config)


if __name__ == '__main__':
    main()