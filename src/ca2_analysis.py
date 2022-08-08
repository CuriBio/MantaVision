import os
import cv2 as cv
import openpyxl
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Tuple, Dict
from track_template import userDrawnROI
from os_functions import contentsOfDir
from waveform_analysis import waveFormAnalysis
from video_api import VideoReader, VideoWriter, supported_file_extensions
from morphology import morphologyMetricsForImage  #, roiInfoFromTemplates
from track_template import trackTemplate, roiInfoFromTemplate
from math import floor
import matplotlib.pyplot as plt
pd.set_option("display.precision", 2)
pd.set_option("display.expand_frame_repr", False)


# TODO: update ca2 analysis function behaviour to accommodate if template paths are filled or not
#       and act accordingly i.e. use the templates if they're provided, or ask the user to draw them otherwise.
# TODO: attempt to determine if the video has low S/N and run morphology in low_s/n mode?
#       it might be enough to just determine if the variance of the first frame is > blah, or
#       even if the mean and median are more than x apart etc
# TODO: add the same contraction vector drop down as is used in tracking
#       this will then determine which edge of the tissue roi is formed by
#       the user drawn or template guided rois

# TODO: while we track the moving roi, we can perform analysis on it
#  to determine the frequency of the contractions, that way we don't need to
#  ask the user to guess. So we need to add the frequency analysis to track template
#  and then report the value in Ca2+ analysis somewhere sensible
#  the way we can tell what the frequency is, is as follows
#  we know physically where the extreme points are (left and right in a purely horizontal contraction)
#  and once we know where those positions are, we can check for frames that have a match
#  that corresponds to being close to those extreme points.
#  we can then pick a physical point mid way between those extreme points
#  and we count how many frames on average between matches to cross that mid point in either direction
#  and of course the time it take is just (num_frames / frames_per_second)
#  and from that, 1/T = frequency

# TODO: check there are no double peaks or troughs.
#       would need to use the peak and trough indices
#       find if the peak or trough is first
#       then oscillate between peaks and troughs ensuring
#       the values in time_stamps of the alternating sequence
#       of peak and trough indices forms a strictly monotonically increase sequence
#       i.e. from one to the next the time is always greater
#       because if there is ever a double peak, the alternating sequence wont be monotonically increasing
#       the alternative is to just throw a tuple ('peak or trough', index) into a list
#       and sort on the index then step through the list and ensure we alternate

# TODO: IF the current method of estimate the various metrics where we
#       use a polynomial fit and then find the roots of the poly FAILS,
#       numpy warnings such as the rank is blah or the fit is ill-conditioned etc ... 
#       we will have to use a method like:
#       take the polynomial coefficients and perform a search of the function
#       between the start and end times
#       could do this with a bisection and stop when the step size < some tol in seconds say 0.001
#       or we could do a multi grid search i.e step size of 0.1 second
#       then 0.01 second between the points we narrowed it to
#       then 0.001 between the subsequent points we narrow it to, etc
#       if the fit still fails, we'd have to find the two closest time points that the
#       metric parameter lives between, and linear interpolate between those two points
#       to find the point the metric is for. this would clearly be inferior since
#       we know most of these signals have exponential or polynomial shape and
#       a linear interpolation between two points (even if close) isn't a great fit.
#       or, 
#       the alternative is to use splines which do essentially the same thing
#       without having to explicitly code the linear interpolation, but since we have to 
#       compute enough splines to for the resolution we need computationally very expensive


def videoROIsInfo(video_path: str, template_info: Dict, reference_roi_captures_tissue: bool) -> Dict:
    """ """
    dynamic_roi_template_image = template_info['dynamic']['image']
    _, dynamic_roi, _, _, _ = trackTemplate(
        input_video_path=video_path,
        template_guide_image_path=None,
        template_rgb=dynamic_roi_template_image,
        max_translation_per_frame=[50, 50]
    )
    num_frames = len(dynamic_roi)
    rois_info = []

    if reference_roi_captures_tissue:
        # we will only be using the reference roi, and, the top, bottom and rhs don't change,
        # only the dynamic roi side changes with the results of the tracking of the dynamic roi
        for frame_num in range(num_frames):
            rois_info.append(
                {
                    'reference': {
                        'x_start': int(dynamic_roi[frame_num]['x_end']),
                        'x_end': template_info['reference']['roi']['x_end'],
                        'y_start': template_info['reference']['roi']['y_start'],
                        'y_end': template_info['reference']['roi']['y_end']
                    }
                }
            )
    else:
        reference_roi_template_image = template_info['reference']['image']
        _, reference_roi, _, _, _ = trackTemplate(
            input_video_path=video_path,
            template_guide_image_path=None,
            template_rgb=reference_roi_template_image,
            max_translation_per_frame=[50, 50]
        )
        for frame_num in range(num_frames):
            rois_info.append(
                {
                    'dynamic': {
                        'x_start': int(dynamic_roi[frame_num]['x_start']),
                        'x_end': int(dynamic_roi[frame_num]['x_end']),
                        'y_start': int(dynamic_roi[frame_num]['y_start']),
                        'y_end': int(dynamic_roi[frame_num]['y_end']),
                    },
                    'reference': {
                        'x_start': int(reference_roi[frame_num]['x_start']),
                        'x_end': int(reference_roi[frame_num]['x_end']),
                        'y_start': int(reference_roi[frame_num]['y_start']),
                        'y_end': int(reference_roi[frame_num]['y_end']),
                    }
                }
            )

    return rois_info


def signalDataFromVideo(
    input_video_path: str,
    output_video_path: str = None,
    low_signal_to_noise: bool = False,
    analysis_method: str = 'None',
    video_roi_info: Dict = None
) -> Dict:
    """ """

    if 'auto' in analysis_method.lower():
        video_rois = iter(
            videoROIsInfo(
                video_path=input_video_path,
                template_info=video_roi_info['template_info'],
                reference_roi_captures_tissue=video_roi_info['reference_roi_captures_tissue']
            )
        )

    input_video_stream = VideoReader(input_video_path)
    if output_video_path is not None:
        video_writer = VideoWriter(
            path=output_video_path,
            width=input_video_stream.frameVideoWidth(),
            height=input_video_stream.frameVideoHeight(),
            time_base=input_video_stream.timeBase(),
            fps=input_video_stream.avgFPS(),
            bitrate=input_video_stream.bitRate()
        )
    else:
        video_writer = None

    time_stamps = []
    signal_values = []
    fg_means = []
    bg_means = []
    while input_video_stream.next():
        time_stamps.append(input_video_stream.timeStamp())
        frame_gray = input_video_stream.frameGray()
        frame_rgb = None
        bg_roi = frame_gray[
            video_roi_info['background']['y_start']:video_roi_info['background']['y_end'],
            video_roi_info['background']['x_start']:video_roi_info['background']['x_end']
        ]
        background_subtractor = np.mean(bg_roi)
        if 'auto' in analysis_method.lower():
            frame_roi_info = next(video_rois)
            if 'morphology' in analysis_method.lower():
                # get the tissue only roi from morphology function
                frame_rgb = input_video_stream.frameVideoRGB()
                frame_rgb, morphology = morphologyMetricsForImage(
                    search_image=frame_rgb,
                    rois_info=frame_roi_info,
                    template_refinement_radius=0,
                    edge_finding_smoothing_radius=10,
                    draw_tissue_roi_only=True,
                    low_signal_to_noise=low_signal_to_noise
                )
                # compute the mean of the tissue only region
                frame_gray_rotated = np.rot90(frame_gray, k=-1)
                frame_gray_rotated_width = frame_gray_rotated.shape[1]
                y_pos = list(
                    range(
                        int(floor(morphology['x_start_position'])),
                        int(floor(morphology['x_end_position'])) + 1
                    )
                )
                x_start_pos = frame_gray_rotated_width - morphology['y_end_positions']
                x_end_pos = frame_gray_rotated_width - morphology['y_start_positions']
                num_rows = len(x_start_pos)
                row_sums = 0
                row_counts = 0
                for row_num in range(0, num_rows):
                    x_row_start = int(x_start_pos[row_num])
                    x_row_end = int(x_end_pos[row_num])
                    if x_row_start == x_row_end:
                        # edge finding failed, so sample from the entire line
                        x_row_start = 0
                        x_row_end = frame_gray_rotated_width
                    frame_row = frame_gray_rotated[y_pos[row_num], x_row_start:x_row_end]
                    row_sums += np.sum(frame_row)
                    row_counts += len(frame_row)
                tissue_mean = float(row_sums)/float(row_counts)
                # append the tissue mean
                fg_means.append(tissue_mean)
                # append the background subtracted tissue mean
                signal_values.append(tissue_mean - background_subtractor)
            else:
                reference_roi = frame_roi_info['reference']
                frame_gray = frame_gray[
                    reference_roi['y_start']:reference_roi['y_end'],
                    reference_roi['x_start']:reference_roi['x_end'],
                ]
        else:  # we're using fixed tissue roi which was already set by the user
            # NOTE: cutting out a new frame_gray from the tissue roi has to go AFTER background roi extraction
            frame_gray = frame_gray[
                video_roi_info['tissue']['y_start']:video_roi_info['tissue']['y_end'],
                video_roi_info['tissue']['x_start']:video_roi_info['tissue']['x_end']
            ]

        # append the background mean
        bg_means.append(background_subtractor)
        # append the mean of the fixed tissue roi for all methods other than fully auto
        if 'morphology' not in analysis_method.lower():
            fg_means.append(np.mean(frame_gray))
            adjusted_frame = frame_gray - background_subtractor
            adjusted_frame[adjusted_frame < 0] = 0
            adjusted_frame_mean = np.mean(adjusted_frame)
            signal_values.append(adjusted_frame_mean)

        # write out the video frame with the rois drawn
        if video_writer is not None:
            # mark the bg (& maybe tissue) ROIs on the output video frame
            if frame_rgb is None:
                frame_rgb = input_video_stream.frameVideoRGB()
            rois_to_draw = [video_roi_info['background']]
            # tissue roi is drawn by the morphology function when auto method is run
            # all other methods require drawing the reference roi here
            if 'auto' not in analysis_method.lower():
                rois_to_draw.append(video_roi_info['tissue'])
            elif 'adjusted' in analysis_method.lower():
                rois_to_draw.append(reference_roi)
            for roi in rois_to_draw:
                top_left_point = (roi['x_start'], roi['y_start'])
                top_right_point = (roi['x_end'], roi['y_start'])
                bottom_left_point = (roi['x_start'], roi['y_end'])
                bottom_right_point = (roi['x_end'], roi['y_end'])
                roi_edges = [
                    (top_left_point, top_right_point),
                    (bottom_left_point, bottom_right_point),
                    (top_left_point, bottom_left_point),
                    (bottom_right_point, top_right_point),
                ]

                roi_outline_colour = (0, 0, 255)
                for edge_point_a, edge_point_b in roi_edges:
                    cv.line(
                        img=frame_rgb,
                        pt1=edge_point_a,
                        pt2=edge_point_b,
                        color=roi_outline_colour,
                        thickness=1,
                        lineType=cv.LINE_AA
                    )
            video_writer.writeFrame(frame_rgb, input_video_stream.frameVideoPTS())

    if video_writer is not None:
        video_writer.close()
    input_video_stream.close()

    return {
        'time_stamps': np.array(time_stamps, dtype=float),
        'signal_values': np.array(signal_values, dtype=float),
        'tissue_means': np.array(fg_means, dtype=float),
        'background_means': np.array(bg_means, dtype=float)
    }


def videoROIs(
    video_file_path: str,
    max_seconds_to_search: float = None,
    roi_method: str = None,
    dynamic_roi_template_path: str = None,
    reference_roi_template_path: str = None,
    background_roi: Dict = None
) -> Dict:
    input_video_stream = VideoReader(video_file_path)
    if not input_video_stream.isOpened():
        return None, None

    input_video_duration = input_video_stream.duration()
    if max_seconds_to_search is None:
        max_seconds_to_search = input_video_duration
    if max_seconds_to_search < 0 or max_seconds_to_search > input_video_duration:
        max_seconds_to_search = input_video_duration

    frame_for_roi_extraction: np.ndarray = None
    frame_for_roi_extraction_gray: np.ndarray = None
    max_intensity_sum: float = 0.0
    while input_video_stream.next():
        frame_gray = input_video_stream.frameGray()
        frame_gray_intensity_sum: float = np.sum(frame_gray)
        if frame_gray_intensity_sum > max_intensity_sum:
            max_intensity_sum = frame_gray_intensity_sum
            frame_for_roi_extraction = input_video_stream.frameRGB()
            frame_for_roi_extraction_gray = frame_gray
        if input_video_stream.timeStamp() > max_seconds_to_search:
            break

    if 'morphology' not in roi_method.lower() and roi_method.lower() != 'none':
        reference_roi_captures_tissue = True
    else:
        reference_roi_captures_tissue = False
    if background_roi is None:
        background_roi = userDrawnROI(frame_for_roi_extraction, "Select the Background ROI")

    rois = {
        'background': background_roi,
        'reference_roi_captures_tissue': reference_roi_captures_tissue
    }
    if 'fixed' in roi_method.lower():
        if reference_roi_template_path is None:
            rois['tissue'] = userDrawnROI(frame_for_roi_extraction, "Select the Tissue/Signal ROI")
        else:
            reference_roi_image = cv.imread(reference_roi_template_path)
            rois['tissue'] = roiInfoFromTemplate(
                frame_for_roi_extraction_gray,
                cv.cvtColor(reference_roi_image, cv.COLOR_BGR2GRAY)
            )
    else:
        rois['tissue'] = {}
        if dynamic_roi_template_path is None:
            dynamic_roi = userDrawnROI(frame_for_roi_extraction, "Select the dynamic ROI to Track")
            dynamic_roi_image = frame_for_roi_extraction[
                dynamic_roi['y_start']:dynamic_roi['y_end'],
                dynamic_roi['x_start']:dynamic_roi['x_end'],
            ]
        else:
            dynamic_roi_image = cv.imread(dynamic_roi_template_path)
            dynamic_roi = roiInfoFromTemplate(
                frame_for_roi_extraction_gray,
                cv.cvtColor(dynamic_roi_image, cv.COLOR_BGR2GRAY)
            )
        if reference_roi_template_path is None:
            if reference_roi_captures_tissue:
                roi_capture_heading = "Select the tissue ROI that will be auto adjusted"
            else:
                roi_capture_heading = "Select the reference ROI to track for morphological analysis"
            reference_roi = userDrawnROI(frame_for_roi_extraction, roi_capture_heading)
            reference_roi_image = frame_for_roi_extraction[
                reference_roi['y_start']:reference_roi['y_end'],
                reference_roi['x_start']:reference_roi['x_end'],
            ]
        else:
            reference_roi_image = cv.imread(reference_roi_template_path)
            reference_roi = roiInfoFromTemplate(
                frame_for_roi_extraction_gray,
                cv.cvtColor(reference_roi_image, cv.COLOR_BGR2GRAY)
            )

        rois['template_info'] = {
            'dynamic': {
                'path': '',
                'image': dynamic_roi_image,
                'roi': dynamic_roi
            },
            'reference': {
                'path': '',
                'image': reference_roi_image,
                'roi': reference_roi
            }
        }

    return rois


def analyzeCa2Data(
    path_to_data: str,
    expected_frequency_hz: float,
    analysis_method: str = 'None',
    low_signal_to_noise: bool = False,
    save_result_plots: bool = False,
    dynamic_roi_template_path: str = None,
    reference_roi_template_path: str = None,
    display_results: bool = False,
    select_background_once: bool = False,
    expected_min_peak_width: int = None,
    expected_min_peak_height: float = None
):
    """ """
    base_dir, files_to_analyze = contentsOfDir(dir_path=path_to_data, search_terms=supported_file_extensions)
    results_dir_name = "results_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    results_dir = os.path.join(base_dir, results_dir_name)
    os.mkdir(results_dir)

    # collect bg & tissue or dynamic & reference ROIs for all the videos to be analyzed, up front, so that
    # all the videos can then be processed automatically without any user interaction
    all_videos_rois = {}
    expected_frequency_hz_tolerance = 0.5
    frequency_epsilon = 0.01
    seconds_for_period = 1.0/(expected_frequency_hz - expected_frequency_hz_tolerance + frequency_epsilon)
    bg_subtraction_roi_search_seconds = 2.0*seconds_for_period
    background_roi = None
    for file_name, file_extension in files_to_analyze:
        input_video_file_path = os.path.join(base_dir, file_name + file_extension)
        video_rois = videoROIs(
            input_video_file_path,
            bg_subtraction_roi_search_seconds,
            analysis_method,
            dynamic_roi_template_path,
            reference_roi_template_path,
            background_roi
        )
        all_videos_rois[file_name] = video_rois
        if select_background_once:
            background_roi = video_rois['background']

    for file_name, file_extension in files_to_analyze:
        if all_videos_rois is None:
            video_roi_info = None
            output_video_path = None
        else:
            video_roi_info = all_videos_rois[file_name]
            if "nd2" in file_extension.lower():
                output_file_extension = ".avi"
            else:
                output_file_extension = file_extension
            output_video_file_name = file_name + "-with_rois" + output_file_extension

            output_video_path = os.path.join(results_dir, output_video_file_name)
        input_video_file_path = os.path.join(base_dir, file_name + file_extension)
        signal_data = signalDataFromVideo(
            input_video_file_path,
            output_video_path,
            low_signal_to_noise,
            analysis_method,
            video_roi_info
        )
        if signal_data is None:
            raise RuntimeError("Error. Signal from video could not be extracted")
        time_stamps = signal_data['time_stamps']
        input_signal = signal_data['signal_values']
        tissue_means = signal_data['tissue_means']
        background_means = signal_data['background_means']

        signal_data_file_path = os.path.join(results_dir, file_name + '-signal_data.xlsx')
        signalDataToCSV(
            time_stamps,
            input_signal,
            tissue_means,
            background_means,
            signal_data_file_path
        )

        signal_data_for_sdk_file_path = os.path.join(results_dir, file_name + '-signal_data_for_sdk.xlsx')
        signalDataForSDK(
            time_stamps,
            input_signal,
            signal_data_for_sdk_file_path
        )

        ca2_analysis = waveFormAnalysis(
            input_signal,
            time_stamps,
            expected_frequency_hz=expected_frequency_hz,
            expected_min_peak_width=expected_min_peak_width,
            expected_min_peak_height=expected_min_peak_height
        )
        if ca2_analysis is None:
            print("\n")
            print("Error. Could not extract sensible peaks/troughs from data")
            print("Was expected frequency set to +/- 0.5Hz of expected frequency?")
            print("No analysis results were written")
            print("\n")
            continue

        path_to_ca2_analysis_results_file = os.path.join(results_dir, file_name + '-results.xlsx')
        ca2AnalysisResultsToCSV(
            ca2_analysis,
            path_to_ca2_analysis_results_file,
        )

        if display_results:
            print()
            print(f'metrics for {file_name}')
            for metrics in ca2_analysis['metrics']:
                p2p_order = metrics['p2p_order']
                average_metrics = metrics['mean_metric_data']
                metric_failure_proportions = metrics['metric_failure_proportions']
                if display_results:
                    print(f'{p2p_order} average metrics (failure %): {average_metrics} ({metric_failure_proportions})')

        if display_results or save_result_plots is not None:
            if save_result_plots:
                plot_file_path = os.path.join(results_dir, file_name + '-plot.png')
            else:
                plot_file_path = None
            plotCa2Signals(
                time_stamps, 
                input_signal,
                ca2_analysis['peak_indices'],
                ca2_analysis['trough_indices'],
                file_name,
                display_results=display_results,
                plot_file_path=plot_file_path
            )


def signalDataFromXLSX(path_to_data: str) -> Tuple[np.ndarray]:
    ''' Reads in an xlsx file containing ca2 experiment data and
        returns a tuple of numpy arrays (time stamps, signal) '''
    ca2_data = pd.read_excel(path_to_data, usecols=[1, 5], dtype=np.float32)
    ca2_data = ca2_data.to_numpy(copy=True).T
    return (ca2_data[0], ca2_data[1])


def plotCa2Signals(
    time_stamps: np.ndarray,
    signal: np.ndarray,
    peak_indices: np.ndarray,
    trough_indices: np.ndarray,
    plot_title: str='',
    plot_smoothed_signal: bool=True,
    display_results: bool = True,
    plot_file_path: str = None
):
    plt.suptitle('Ca2+ Activity')
    plt.title(plot_title)
    if plot_smoothed_signal:
        plt.plot(time_stamps, signal)
    else:
        plt.scatter(time_stamps, signal, s=2, facecolors='none', edgecolors='g')
    plt.scatter(time_stamps[peak_indices], signal[peak_indices], s=80, facecolors='none', edgecolors='b')
    plt.scatter(time_stamps[trough_indices], signal[trough_indices], s=80, facecolors='none', edgecolors='r')
    plt.xlabel('Time (seconds)')
    plt.ylabel('Signal Intensity')
    if plot_file_path is not None:
        # NOTE: do NOT change the order of saving and showing the plot
        #       the plots will be blank if save is not performed first
        plt.savefig(fname=plot_file_path, format='png', facecolor='white', transparent=False, dpi=400.0)
    if display_results:
        plt.show()
    plt.close()


def signalDataForSDK(
    time_stamps: np.ndarray,
    input_signal: np.ndarray,
    sdk_signal_data_file_path: str
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # set the column headings
    heading_row = 1
    time_stamp_column = 1
    sheet.cell(heading_row, time_stamp_column).value = 'time'
    signal_value_column = time_stamp_column + 1
    sheet.cell(heading_row, signal_value_column).value = 'signal'

    # set the data values
    data_row_start = heading_row + 1
    num_data_points = len(time_stamps)
    for data_point_index in range(0, num_data_points):
        row_num = data_point_index + data_row_start
        sheet.cell(row_num, time_stamp_column).value = time_stamps[data_point_index]
        sheet.cell(row_num, signal_value_column).value = input_signal[data_point_index]

    workbook.save(filename=sdk_signal_data_file_path)


def signalDataToCSV(
    time_stamps: np.ndarray,
    input_signal: np.ndarray,
    tissue_means: np.ndarray,
    background_means: np.ndarray,
    output_data_file_path: str
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # set the column headings
    heading_row = 1
    frame_number_column = 1
    sheet.cell(heading_row, frame_number_column).value = 'frame number'
    time_stamp_column = frame_number_column + 1
    sheet.cell(heading_row, time_stamp_column).value = 'time'
    signal_value_column = time_stamp_column + 1
    sheet.cell(heading_row, signal_value_column).value = 'signal'
    fg_mean_value_column = signal_value_column + 1
    sheet.cell(heading_row, fg_mean_value_column).value = 'fg_mean'
    bg_mean_value_column = fg_mean_value_column + 1
    sheet.cell(heading_row, bg_mean_value_column).value = 'bg_mean'

    # set the data values
    data_row_start = heading_row + 1
    num_data_points = len(time_stamps)
    for data_point_index in range(0, num_data_points):
        row_num = data_point_index + data_row_start
        sheet.cell(row_num, frame_number_column).value = data_point_index
        sheet.cell(row_num, time_stamp_column).value = time_stamps[data_point_index]
        sheet.cell(row_num, signal_value_column).value = input_signal[data_point_index]
        sheet.cell(row_num, fg_mean_value_column).value = tissue_means[data_point_index]
        sheet.cell(row_num, bg_mean_value_column).value = background_means[data_point_index]

    workbook.save(filename=output_data_file_path)


def ca2AnalysisResultsToCSV(
        ca2_analysis: Dict,
        path_to_results_file: str,
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # set the column headings
    heading_row = 1
    metric_type_column = 1
    sheet.cell(heading_row, metric_type_column).value = 'metric type'
    metric_value_column = metric_type_column + 1
    sheet.cell(heading_row, metric_value_column).value = 'metric average'
    normalized_metric_value_column = metric_value_column + 1
    sheet.cell(heading_row, normalized_metric_value_column).value = 'normalized metric average'
    num_points_column = normalized_metric_value_column + 1
    sheet.cell(heading_row, num_points_column).value = 'num points'
    num_failed_points_column = num_points_column + 1
    sheet.cell(heading_row, num_failed_points_column).value = 'num failed points'
    percent_failed_points_column = num_failed_points_column + 1
    sheet.cell(heading_row, percent_failed_points_column).value = '% failed points'

    # set the column values for each metric
    row_num = heading_row + 1
    num_p2p_types = len(ca2_analysis['metrics'])
    for p2p_type_num in range(num_p2p_types):
        metrics = ca2_analysis['metrics'][p2p_type_num]
        metric_labels = metrics['metrics_labels']
        metric_values = metrics['mean_metric_data']
        normalized_metric_values = metrics['normalized_metric_data']
        num_points = metrics['num_metric_points']
        num_failed_points = metrics['num_metric_failures']
        failure_percentages = metrics['metric_failure_proportions']

        num_metrics = len(metric_values)
        for metric_num in range(num_metrics):
            sheet.cell(row_num, metric_type_column).value = metric_labels[metric_num]
            sheet.cell(row_num, metric_value_column).value = metric_values[metric_num]
            sheet.cell(row_num, normalized_metric_value_column).value = normalized_metric_values[metric_num]
            sheet.cell(row_num, num_points_column).value = num_points
            sheet.cell(row_num, num_failed_points_column).value = num_failed_points[metric_num]
            sheet.cell(row_num, percent_failed_points_column).value = failure_percentages[metric_num]
            row_num += 1

    # add a measure of average frequency
    row_num += 1
    avg_frequency = ca2_analysis['avg_frequency']
    sheet.cell(row_num, metric_type_column).value = 'frequency'
    sheet.cell(row_num, metric_value_column).value = avg_frequency

    workbook.save(filename=path_to_results_file)


def dataMode(input_data: np.ndarray) -> float:
    data_min = np.floor(np.min(input_data))
    data_max = np.ceil(np.max(input_data))
    histogram_range = range(int(data_min) + 1, int(data_max) + 1)
    data_histogram = np.histogram(input_data, bins=histogram_range, range=(data_min, data_max))[0]
    histogram_peak = np.argmax(data_histogram)
    return data_min + histogram_peak


if __name__ == '__main__':
    analyzeCa2Data(
        path_to_data='select_dir',
        expected_frequency_hz=1.5,
        analysis_method=None,
        save_result_plots=True,
        display_results=False
    )
