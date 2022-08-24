from typing import Tuple, Dict, List
import openpyxl
import numpy as np


# TODO: for user xlsx functions, create a class that creates an xlsx file,
#       then has methods to add sheets which perform the functions currently being used to produce xlsx files
#       with appropriate names that add all the existing sheet info
#       that way we can create an xlsx object and pass in the data
#       then successively call various methods that each add their own sheets,
#       take from the input data what they need and then there would be a write method of course.

# TODO: in the user xlsx, have separate columns for pixel & microns displacement


def ca2SignalDataToXLSXforSDK(
    time_stamps: np.ndarray,
    input_signal: np.ndarray,
    sdk_signal_data_file_path: str,
    video_meta_data: Dict
):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # set the column headings
    heading_row = 1
    time_stamp_column = 1
    sheet.cell(heading_row, time_stamp_column).value = 'time'
    signal_value_column = time_stamp_column + 1
    sheet.cell(heading_row, signal_value_column).value = 'signal'

    # set the data values
    data_row_start = heading_row + 1
    num_data_points = len(time_stamps)
    for data_point_index in range(0, num_data_points):
        row_num = data_point_index + data_row_start
        sheet.cell(row_num, time_stamp_column).value = time_stamps[data_point_index]
        sheet.cell(row_num, signal_value_column).value = input_signal[data_point_index]

    # add meta data
    sheet['E2'] = video_meta_data['well_name']
    sheet['E3'] = video_meta_data['video_date'] + ' 00:00:00'
    sheet['E4'] = 'NA'  # plate barcode
    sheet['E5'] = video_meta_data['frames_per_second']
    sheet['E6'] = 'y'  # do twitch's point up
    sheet['E7'] = 'NA'  # microscope name

    workbook.save(filename=sdk_signal_data_file_path)


def ca2AnalysisToXLSX(
    time_stamps: np.ndarray,
    input_signal: np.ndarray,
    tissue_means: np.ndarray,
    background_means: np.ndarray,
    ca2_analysis: Dict,
    output_file_path: str
):
    """ """

    workbook = openpyxl.Workbook()

    # create the first worksheet with the raw signal data
    sheet = workbook.active
    sheet.title = 'Signal Data'

    # set the column headings
    heading_row = 1
    frame_number_column = 1
    sheet.cell(heading_row, frame_number_column).value = 'frame number'
    time_stamp_column = frame_number_column + 1
    sheet.cell(heading_row, time_stamp_column).value = 'time'
    signal_value_column = time_stamp_column + 1
    sheet.cell(heading_row, signal_value_column).value = 'signal'
    fg_mean_value_column = signal_value_column + 1
    sheet.cell(heading_row, fg_mean_value_column).value = 'fg_mean'
    bg_mean_value_column = fg_mean_value_column + 1
    sheet.cell(heading_row, bg_mean_value_column).value = 'bg_mean'

    # set the data values
    data_row_start = heading_row + 1
    num_data_points = len(time_stamps)
    for data_point_index in range(0, num_data_points):
        row_num = data_point_index + data_row_start
        sheet.cell(row_num, frame_number_column).value = data_point_index
        sheet.cell(row_num, time_stamp_column).value = time_stamps[data_point_index]
        sheet.cell(row_num, signal_value_column).value = input_signal[data_point_index]
        sheet.cell(row_num, fg_mean_value_column).value = tissue_means[data_point_index]
        sheet.cell(row_num, bg_mean_value_column).value = background_means[data_point_index]

    # create a second worksheet and write out the analysis results
    sheet = workbook.create_sheet('Ca2+ Analysis')

    # set the column headings
    heading_row = 1
    metric_type_column = 1
    sheet.cell(heading_row, metric_type_column).value = 'metric type'
    metric_value_column = metric_type_column + 1
    sheet.cell(heading_row, metric_value_column).value = 'metric average'
    normalized_metric_value_column = metric_value_column + 1
    sheet.cell(heading_row, normalized_metric_value_column).value = 'normalized metric average'
    num_points_column = normalized_metric_value_column + 1
    sheet.cell(heading_row, num_points_column).value = 'num points'
    num_failed_points_column = num_points_column + 1
    sheet.cell(heading_row, num_failed_points_column).value = 'num failed points'
    percent_failed_points_column = num_failed_points_column + 1
    sheet.cell(heading_row, percent_failed_points_column).value = '% failed points'

    # set the column values for each metric
    row_num = heading_row + 1
    num_p2p_types = len(ca2_analysis['metrics'])
    for p2p_type_num in range(num_p2p_types):
        metrics = ca2_analysis['metrics'][p2p_type_num]
        metric_labels = metrics['metrics_labels']
        metric_values = metrics['mean_metric_data']
        normalized_metric_values = metrics['normalized_metric_data']
        num_points = metrics['num_metric_points']
        num_failed_points = metrics['num_metric_failures']
        failure_percentages = metrics['metric_failure_proportions']

        num_metrics = len(metric_values)
        for metric_num in range(num_metrics):
            sheet.cell(row_num, metric_type_column).value = metric_labels[metric_num]
            sheet.cell(row_num, metric_value_column).value = metric_values[metric_num]
            sheet.cell(row_num, normalized_metric_value_column).value = normalized_metric_values[metric_num]
            sheet.cell(row_num, num_points_column).value = num_points
            sheet.cell(row_num, num_failed_points_column).value = num_failed_points[metric_num]
            sheet.cell(row_num, percent_failed_points_column).value = failure_percentages[metric_num]
            row_num += 1

    # add a measure of average frequency
    row_num += 1
    avg_frequency = ca2_analysis['avg_frequency']
    sheet.cell(row_num, metric_type_column).value = 'frequency'
    sheet.cell(row_num, metric_value_column).value = avg_frequency

    workbook.save(filename=output_file_path)


def trackingResultsToXLSXforSDK(
    tracking_results: List[Dict],
    meta_data: Dict,
    path_to_output_file
):
    workbook = openpyxl.Workbook()  # open a blank workbook
    sheet = workbook.active

    # add meta data
    well_name = meta_data['Well Name']
    if well_name is None:
        well_name = 'Z01'
    sheet['E2'] = well_name
    date_stamp = meta_data['Date Stamp']
    sheet['E3'] = date_stamp + ' 00:00:00'
    sheet['E4'] = 'NA'  # plate barcode
    frames_per_second = meta_data['Frames Per Second']
    sheet['E5'] = frames_per_second
    sheet['E6'] = 'y'  # do twitch's point up
    sheet['E7'] = 'NA'  # microscope name

    # add runtime data (time, displacement etc)
    template_start_row = 2
    time_column = 'A'
    displacement_column = 'B'
    num_rows_to_write = len(tracking_results)
    for results_row in range(num_rows_to_write):
        tracking_result = tracking_results[results_row]
        sheet_row = str(results_row + template_start_row)
        sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
        sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
    workbook.save(filename=path_to_output_file)


def trackingResultsToXSLX(tracking_results: List[Dict], meta_data: Dict, path_to_output_file: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # add meta data
    meta_data_name_column = 'G'
    meta_data_value_column = 'H'
    meta_data_row_number = 1
    sheet[meta_data_value_column + str(meta_data_row_number)] = 'Runtime Meta Data'
    for column_name, column_value in meta_data.items():
        meta_data_row_number += 1
        sheet[meta_data_name_column + str(meta_data_row_number)] = column_name
        if isinstance(column_value, (int, float)):
            sheet[meta_data_value_column + str(meta_data_row_number)] = column_value
        else:
            sheet[meta_data_value_column + str(meta_data_row_number)] = f"{column_value}"

    # add runtime data
    # heading fields
    heading_row = 1
    time_column = 'A'
    sheet[time_column + str(heading_row)] = 'Time (s)'
    displacement_column = 'B'
    sheet[displacement_column + str(heading_row)] = 'Displacement From Min'
    x_pos_column = 'C'
    sheet[x_pos_column + str(heading_row)] = 'Template Match Center X (pixel pos)'
    y_pos_column = 'D'
    sheet[y_pos_column + str(heading_row)] = 'Template Match Center Y (pixel pos)'
    angle_column = 'E'
    sheet[angle_column + str(heading_row)] = 'Template Match Angle (deg)'

    # time, displacement from ref position, absolute position and angle fields
    data_row = heading_row + 1
    num_rows_to_write = len(tracking_results)
    for results_row in range(num_rows_to_write):
        tracking_result = tracking_results[results_row]
        sheet_row = str(results_row + data_row)
        sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
        sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
        sheet[x_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_X'])
        sheet[y_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_Y'])
        sheet[angle_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ROTATION'])
    workbook.save(filename=path_to_output_file)


def metadataRequiredForXLSX(
    well_name: str = None,
    date_stamp: str = None,
    frames_per_second: float = None,
    user_roi_selection: bool = None,
    max_translation_per_frame: int = None,
    max_rotation_per_frame: float = None,
    contraction_vector: Tuple[int, int] = None,
    microns_per_pixel: float = None,
    output_conversion_factor: float = None,
    sub_pixel_search_increment: float = None,
    sub_pixel_refinement_radius: int = None,
    estimated_frequency: float = None,
) -> Dict:
    meta_data = {
        'Well Name': well_name,
        'Date Stamp': date_stamp,
        'Frames Per Second': frames_per_second,
        'User ROI Selection': user_roi_selection,
        'Max Translation Per Frame': max_translation_per_frame,
        'Max Rotation Per Frame': max_rotation_per_frame,
        'Contraction Vector': contraction_vector,
        'Microns Per Pixel': microns_per_pixel,
        'Output Conversion Factor': output_conversion_factor,
        'Sub Pixel Search Increment': sub_pixel_search_increment,
        'Sub Pixel Refinement Radius': sub_pixel_refinement_radius,
        'Estimated Frequency': estimated_frequency
    }
    # fill in missing values in meta_data
    for md_key, md_value in meta_data.items():
        if md_value is None:
            meta_data[md_key] = 'None'
    return meta_data
