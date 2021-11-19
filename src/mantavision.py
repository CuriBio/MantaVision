#! /usr/bin/env python

import argparse
import os
import glob
import shutil
import sys
import json
import pathlib
import time
from typing import Tuple, List, Dict
import zipfile
import openpyxl # pip install --user openpyxl
from cv2 import cv2 as cv  # pip install --user opencv-python
from datetime import datetime
from tkinter import Tk as tk
from tkinter.filedialog import askopenfilename, askdirectory
from video2images import video2images
from track_template import trackTemplate


def runTrackTemplate(config: Dict):
  track_templates_start_time = time.time()
  dirs, args = verifiedInputs(config)
  
  # make all the dirs needed for writing the results 
  # unless they already exist in which case we need to barf
  dirs_exist_error_message = ''
  if os.path.isdir(dirs['results_dir_path']):
    dirs_exist_error_message += "results dir already exists. Cannot overwrite.\n"
  if os.path.isdir(dirs['results_json_dir_path']):
    dirs_exist_error_message += "json results dir already exists. Cannot overwrite.\n"
  if os.path.isdir(dirs['results_xlsx_dir_path']):
    dirs_exist_error_message += "xlsx results dir already exists. Cannot overwrite.\n"
  if os.path.isdir(dirs['results_video_dir_path']):
    dirs_exist_error_message += "video results dir already exists. Cannot overwrite.\n"
  if os.path.isdir(dirs['results_template_dir_path']):
    dirs_exist_error_message += "template results dir already exists. Cannot overwrite.\n"    
  if len(dirs_exist_error_message) > 0:
    dirs_exist_error_message = "ERROR.\n" + dirs_exist_error_message + "Nothing Tracked."
    print(dirs_exist_error_message)
    sys.exit(1)
  os.mkdir(dirs['results_dir_path'])
  os.mkdir(dirs['results_json_dir_path'])
  os.mkdir(dirs['results_xlsx_dir_path'])
  os.mkdir(dirs['results_video_dir_path'])
  os.mkdir(dirs['results_template_dir_path'])
  os.mkdir(dirs['results_video_min_frames_dir_path'])
  if next(iter(args))['output_frames']:
    os.mkdir(dirs['results_video_frames_dir_path'])

  # run the tracking routine on each input video
  # and write out the results
  print("\nTemplate Tracker running...") 
  total_tracking_time = 0
  for input_args in args:
    print(f'processing: {input_args["input_video_path"]}')
    video_tracking_start_time = time.time()
    messages, tracking_results, frames_per_second, template, min_frame_number = trackTemplate(
      input_args['input_video_path'],
      input_args['template_guide_image_path'],
      input_args['output_video_path'],
      input_args['guide_match_search_seconds'],
      input_args['microns_per_pixel'],
      input_args['output_conversion_factor'],
      input_args['sub_pixel_search_increment'],
      input_args['sub_pixel_refinement_radius'],
      input_args['user_roi_selection'],
      input_args['max_pixel_movement_per_frame']
    )
    total_tracking_time += (time.time() - video_tracking_start_time)
    
    warning_msg, error_msg = messages

    if warning_msg is not None:
      print(warning_msg)

    if error_msg is not None:
      print(error_msg)
      sys.exit(1)

    # write the template used for tracking to the results dir
    cv.imwrite(input_args['results_template_filename'], template)

    # write out the frame with the min movement position
    video2images(
      input_video_path=input_args['input_video_path'],
      output_dir_path=dirs['results_video_min_frames_dir_path'],
      enhance_contrast=False,
      frame_number_to_write=min_frame_number,
      image_extension='tiff'
    )

    # write out the results video as frames if requested
    if input_args['output_frames']:
      os.mkdir(input_args['output_video_frames_dir_path'])      
      video2images(
        input_video_path=input_args['output_video_path'],
        output_dir_path=input_args['output_video_frames_dir_path'],
        enhance_contrast=False,
        image_extension='jpg'  # don't need high quality images for this
      )

    # write the results as xlsx
    resultsToCSVforSDK(
      tracking_results,
      input_args['path_to_excel_template'],
      input_args['path_to_sdk_excel_results'],
      frames_per_second,
      input_args['well_name'], 
      input_args['date_stamp']
    )

    resultsToCSVforUser(
      tracking_results,
      input_args['path_to_user_excel_results'],
      frames_per_second,
      input_args['well_name'], 
      input_args['date_stamp']
    )

    # write the run config and results as json
    if input_args['output_json_path'] is not None:
      tracking_results_complete = {
        "INPUT_ARGS": input_args,
        "ERROR_MSGS": error_msg,
        "WARNING_MSGS": warning_msg,
        "RESULTS": tracking_results
      }
      with open(input_args['output_json_path'], 'w') as outfile:
        json.dump(tracking_results_complete, outfile, indent=4)

  # create a zip archive and write all the xlsx files to it
  xlsx_archive_file_path = os.path.join(dirs['results_dir_path'], 'xlsx-results.zip')
  xlsx_archive = zipfile.ZipFile(xlsx_archive_file_path, 'w')
  for dir_name, _, file_names in os.walk(dirs['results_xlsx_dir_path']):
    for file_name in file_names:
      if 'sdk' in file_name:  # only include the xlsx files intended for the sdk
        file_path = os.path.join(dir_name, file_name)
        xlsx_archive.write(file_path, os.path.basename(file_path))
  xlsx_archive.close()

  num_videos_processed = len(args)
  track_templates_runtime = time.time() - track_templates_start_time
  per_video_tracking_time: float = float(total_tracking_time) / float(num_videos_processed)
  print(f'...Template Tracker completed in {round(track_templates_runtime, 2)}s')
  print(f'\nActual tracking time for {num_videos_processed} videos: {round(total_tracking_time, 2)}s ({round(per_video_tracking_time, 2)}s per video)')


def verifiedInputs(config: Dict) -> Tuple[str, List[Dict]]:
  '''
  '''
  error_msgs = []
  # check the dir path to input videos
  open_video_dir_dialog = False
  if 'input_video_path' in config:
    if config['input_video_path'] is None:
      open_video_dir_dialog = True
    else:
      if config['input_video_path'].lower() == 'select_dir':
        open_video_dir_dialog = True
      elif not os.path.isdir(config['input_video_path']):
        error_msgs.append('Input path to video/s does not exist.')      
  else:
      open_video_dir_dialog = True
  # pop up a dialog to select the dir for videos if required
  if open_video_dir_dialog:
    print()
    print("waiting for user input (video dir) via pop up dialog box...")
    dir_path_via_gui = getDirPathViaGUI(window_title='Select Directory With Videos To Track')
    if dir_path_via_gui == () or dir_path_via_gui == '':
      error_msgs.append('No input path to video/s was provided.')
    else:
      config['input_video_path'] = dir_path_via_gui
    print("...user input obtained from pop up dialog box.")

  # check the file path to the template image
  user_roi_selection = False
  open_template_dir_dialog = False
  if 'template_guide_image_path' in config:
    if config['template_guide_image_path'] is None:
      open_template_dir_dialog = True
    else:
      if config['template_guide_image_path'].lower() == 'select_file':
        open_template_dir_dialog = True
      elif config['template_guide_image_path'].lower() == 'draw':
        config['template_guide_image_path'] = ''
        user_roi_selection = True
      elif not os.path.isfile(config['template_guide_image_path']):
        error_msgs.append('Input template image file does not exist.')      
  else:
      open_template_dir_dialog = True
  # pop up a dialog to select the template file if required
  if open_template_dir_dialog:
    print()
    print("waiting for user input (template to use) via pop up dialog box...")    
    file_path_via_gui = getFilePathViaGUI(window_title='Select File With Template To Track')
    if file_path_via_gui == () or file_path_via_gui == '':
      error_msgs.append('No input template image path was provided.')
    else:
      config['template_guide_image_path'] = file_path_via_gui
    print("...user input obtained from pop up dialog box.")

  # barf if there was an error with either the input video dir path or template file path
  if len(error_msgs) > 0:
    error_msg = 'ERROR.'
    for error_string in error_msgs:
      error_msg = error_msg + ' ' + error_string
    error_msg += ' Nothing to do. Exiting.'
    print(error_msg)
    sys.exit(1)
  
  template_guide_image_path = config['template_guide_image_path']
  supported_file_extensions = ['.mp4', '.avi', '.mov', '.nd2']
  base_dir, video_files = contentsOfDir(dir_path=config['input_video_path'], search_terms=supported_file_extensions)
  
  unique_name = "results_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
  results_dir_path = os.path.join(base_dir, unique_name)
  results_json_dir_path = os.path.join(results_dir_path, 'json')
  results_xlsx_dir_path = os.path.join(results_dir_path, 'xlsx')
  results_video_dir_path = os.path.join(results_dir_path, 'video')
  results_template_dir_path = os.path.join(results_dir_path, 'template')  
  results_video_frames_dir_path = os.path.join(results_video_dir_path, 'frames')
  results_video_min_frames_dir_path = os.path.join(results_dir_path, 'min_frame')

  dirs = {
    'base_dir': base_dir,
    'results_dir_path': results_dir_path,
    'results_json_dir_path': results_json_dir_path,
    'results_xlsx_dir_path': results_xlsx_dir_path,
    'results_video_dir_path': results_video_dir_path,
    'results_template_dir_path': results_template_dir_path,
    'results_video_frames_dir_path': results_video_frames_dir_path,
    'results_video_min_frames_dir_path': results_video_min_frames_dir_path
  }

  if 'path_to_excel_template' not in config:
    path_to_excel_template = None
  else:
    path_to_excel_template = config['path_to_excel_template']

  if 'guide_match_search_seconds' not in config:
    guide_match_search_seconds = None
  else:
    guide_match_search_seconds = config['guide_match_search_seconds']
  
  if 'microns_per_pixel' not in config:
    microns_per_pixel = None
  else:
    microns_per_pixel = config['microns_per_pixel']

  if 'output_conversion_factor' not in config:
    output_conversion_factor = None
  else:
    output_conversion_factor = config['output_conversion_factor']

  if 'sub_pixel_search_increment' not in config:
    sub_pixel_search_increment = None
  else:
    sub_pixel_search_increment = config['sub_pixel_search_increment']

  if 'sub_pixel_refinement_radius' not in config:
    sub_pixel_refinement_radius = None
  else:
    sub_pixel_refinement_radius = config['sub_pixel_refinement_radius']
  if sub_pixel_search_increment is None and sub_pixel_refinement_radius is not None:
    print('WARNING. sub_pixel_refinement_radius ignored because sub_pixel_search_increment not provided')
    sub_pixel_refinement_radius = None
  
  if 'max_pixel_movement_per_frame' not in config:
    max_pixel_movement_per_frame = None
  else:
    max_pixel_movement_per_frame = (config['max_pixel_movement_per_frame'], config['max_pixel_movement_per_frame'])
  
  if 'output_frames' not in config:
    output_frames = False
  else:
    output_frames = config['output_frames']
  
  # set all the values needed to run template matching on each input video
  configs = []
  for file_name, input_file_extension in video_files:

    # check the file name conforms to minimum requirements
    num_chars_in_datestamp = len('yyyy-mm-dd')
    num_chars_in_wellname = len('A001')
    min_num_chars_in_file_name = num_chars_in_datestamp + num_chars_in_wellname
    if len(file_name) < min_num_chars_in_file_name:
      print(f'ERROR. the input video {file_name} does not have a valid name.')
      print('The pattern must have a datestamp as the first 10 characters and a wellname as the last 4 characters')
      print('i.e. 1010-01-01 any other characters A001')
      sys.exit(1)

    # make sure a valid well name can be extracted from the file name
    file_name_length = len(file_name)
    well_name = file_name[file_name_length - num_chars_in_wellname:]
    well_name_valid = True
    well_name_letter_part = well_name[0]
    if not well_name_letter_part.isalpha():
      well_name_valid = False    
    well_name_number_part = well_name[1:]
    if not well_name_number_part.isdigit():
      well_name_valid = False
    if not well_name_valid:
      print("ERROR. An input video file does not contain a valid well name as expected as the last word.")
      print(f"The last word of the filename is: {well_name}")
      print("The last word of the file name must be a letter followed by a zero padded 3 digit number i.e. A001 or D006")
      sys.exit(1)

    # make sure a valid date stamp can be extacted from the file name
    date_stamp = file_name[:num_chars_in_datestamp]
    date_stamp_parsed = date_stamp.split("-")
    num_parts_in_date = 3
    date_stamp_valid = True
    if len(date_stamp_parsed) != num_parts_in_date:
      date_stamp_valid = False
    else:
      for date_part in date_stamp_parsed:
        if not date_part.isdigit():
          date_stamp_valid = False
    if not date_stamp_valid:
      print("ERROR. An input video file does not contain a valid date stamp as expected for the first word.")
      print(f"The first word of the filename is: {date_stamp}")
      print("The first word of the file name must be of the form yyyy-mm-dd i.e. 1010-01-01")
      sys.exit(1)

    # set all the required path values
    input_video_path = os.path.join(base_dir, file_name + input_file_extension)
    output_video_frames_dir_path = os.path.join(results_video_frames_dir_path, file_name)
    output_json_path = os.path.join(results_json_dir_path, file_name + '-results.json')
    path_to_user_excel_results = os.path.join(results_xlsx_dir_path, file_name + '-reslts_user.xlsx')
    path_to_sdk_excel_results = os.path.join(results_xlsx_dir_path, file_name + '-reslts_sdk.xlsx')
    output_file_extension = '.mp4'
    output_video_path = os.path.join(results_video_dir_path, file_name + '-results' + output_file_extension)
    results_template_filename = os.path.join(results_template_dir_path, file_name + '-template.tiff')

    configs.append({
      'input_video_path': input_video_path,
      'template_guide_image_path': template_guide_image_path,
      'results_template_filename': results_template_filename,
      'user_roi_selection': user_roi_selection,
      'output_video_path': output_video_path,
      'output_video_frames_dir_path': output_video_frames_dir_path,
      'output_json_path': output_json_path,
      'path_to_excel_template': path_to_excel_template,
      'path_to_sdk_excel_results': path_to_sdk_excel_results,
      'path_to_user_excel_results': path_to_user_excel_results,
      'guide_match_search_seconds': guide_match_search_seconds,
      'microns_per_pixel': microns_per_pixel,
      'output_conversion_factor': output_conversion_factor,
      'sub_pixel_search_increment': sub_pixel_search_increment,
      'sub_pixel_refinement_radius': sub_pixel_refinement_radius,
      'max_pixel_movement_per_frame': max_pixel_movement_per_frame,
      'output_frames': output_frames,
      'well_name': well_name,
      'date_stamp': date_stamp,
    })

  return (dirs, configs)



def getDirPathViaGUI(window_title: str='') -> str:
  # show an "Open" dialog box and return the path to the selected dir
  window=tk()
  window.withdraw()
  window.lift()
  window.overrideredirect(True)
  window.call('wm', 'attributes', '.', '-topmost', True)
  return askdirectory(
    initialdir='./',
    title=window_title
  )


def getFilePathViaGUI(window_title: str='') -> str:
  # show an "Open" dialog box and return the path to the selected file
  window=tk()
  window.withdraw()
  window.lift()  
  window.overrideredirect(True)
  window.call('wm', 'attributes', '.', '-topmost', True)
  return askopenfilename(
    initialdir='./',
    title=window_title    
  ) 


def contentsOfDir(dir_path: str, search_terms: List[str], search_extension_only: bool=True) -> Tuple[List[str], List[Tuple[str]]]:
  all_files_found = []  
  if os.path.isdir(dir_path):
    base_dir = dir_path
    for search_term in search_terms:
      glob_search_term = '*' + search_term
      if not search_extension_only:
        glob_search_term += '*'
      files_found = glob.glob(os.path.join(dir_path, glob_search_term))
      if len(files_found) > 0:
        all_files_found.extend(files_found)
  else:
    # presume it's actually a single file path
    base_dir = os.path.dirname(dir_path)
    all_files_found = [dir_path]
  if len(all_files_found) < 1:
    return None, None

  files = []
  for file_path in all_files_found:
      file_name, file_extension = os.path.splitext(os.path.basename(file_path))
      files.append((file_name, file_extension))
  return base_dir, files


def resultsToCSVforSDK(
  tracking_results: List[Dict],
  path_to_template_file: str,
  path_to_output_file,
  frames_per_second: float,
  well_name: str = None, 
  date_stamp: str = '1010-01-01'
):    
  if path_to_template_file is None:
    workbook = openpyxl.Workbook()  # open a blank workbook
  else:  # open the template workbook
    shutil.copyfile(path_to_template_file, path_to_output_file)
    workbook = openpyxl.load_workbook(filename=path_to_output_file)
  sheet = workbook.active

  if well_name is None:
    well_name = 'Z01'
  sheet['E2'] = well_name
  sheet['E3'] = date_stamp + ' 00:00:00'
  sheet['E4'] = 'NA'  # plate barcode
  sheet['E5'] = frames_per_second
  sheet['E6'] = 'y'   # do twiches point up
  sheet['E7'] = 'NA'  # microscope name

  # set the time and post displacement fields
  template_start_row = 2
  time_column = 'A'
  displacement_column = 'B'
  num_rows_to_write = len(tracking_results)
  for results_row in range(num_rows_to_write):
      tracking_result = tracking_results[results_row]
      sheet_row = str(results_row + template_start_row)
      sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
      sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
  workbook.save(filename=path_to_output_file)


def resultsToCSVforUser(
  tracking_results: List[Dict],
  path_to_output_file,
  frames_per_second: float,
  well_name: str = None, 
  date_stamp: str = '1010-01-01'
):    
  workbook = openpyxl.Workbook()
  sheet = workbook.active

  if well_name is None:
    well_name = 'Z01'
  sheet['F2'] = 'Well Name'    
  sheet['G2'] = well_name
  sheet['F3'] = 'Date'
  sheet['G3'] = date_stamp + ' 00:00:00'
  sheet['F4'] = 'Plate Barcode'
  sheet['G4'] = 'NA'  # plate barcode
  sheet['F5'] = 'GPS'
  sheet['G5'] = frames_per_second
  sheet['F6'] = 'Twitches Point Up' 
  sheet['G6'] = 'y'   # do twiches point up
  sheet['F7'] = 'Microscope Name'
  sheet['G7'] = 'NA'  # microscope name

  time_column = 'A'
  displacement_column = 'B'
  x_pos_column = 'C'
  y_pos_column = 'D'
  heading_row = 1
  data_row = 2

  sheet[time_column + str(heading_row)] = 'Time'
  sheet[displacement_column + str(heading_row)] = 'XY Displacement'
  sheet[x_pos_column + str(heading_row)] = 'Template Match X Pos'
  sheet[y_pos_column + str(heading_row)] = 'Template Match Y Pos'

  # set the time and post displacement fields
  num_rows_to_write = len(tracking_results)
  for results_row in range(num_rows_to_write):
      tracking_result = tracking_results[results_row]
      sheet_row = str(results_row + data_row)
      sheet[time_column + sheet_row] = float(tracking_result['TIME_STAMP'])
      sheet[displacement_column + sheet_row] = float(tracking_result['XY_DISPLACEMENT'])
      sheet[x_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_X'])
      sheet[y_pos_column + sheet_row] = float(tracking_result['TEMPLATE_MATCH_ORIGIN_Y'])
  workbook.save(filename=path_to_output_file)


def config_from_json(json_config_path) -> Tuple[str, List[Dict]]:
  json_file = open(json_config_path)
  config = json.load(json_file)
  return config


def config_from_cmdline(cmdline_args) -> dict:
  config = {}
  config['input_video_path'] = cmdline_args.input_video_path
  config['template_guide_image_path'] = cmdline_args.template_guide_image_path
  config['output_path'] = cmdline_args.output_video_path
  config['guide_match_search_seconds'] = cmdline_args.guide_match_search_seconds
  config['microns_per_pixel'] = cmdline_args.microns_per_pixel
  config['path_to_excel_template'] = cmdline_args.path_to_excel_template  
  return config


if __name__ == '__main__':
  # os.path.expanduser('~')
  # home_dir = pathlib.Path.home()

  # read in a config file & parse the input args
  parser = argparse.ArgumentParser(
      description='Tracks a template image through each frame of a video.',
  )
  parser.add_argument(
      '--input_video_path',
      default=None,
      help='Path of input video/s to track a template.',
  )
  parser.add_argument(
      '--template_guide_image_path',
      default=None,
      help='Path to an image that will be used as a template to match.',
  )
  parser.add_argument(
      '--path_to_excel_template',
      default=None,
      help='path to exel spread sheet used as a template to write the results into',
  )
  parser.add_argument(
    '--json_config_path',
    default=None,
    help='Path of a json file with run config parameters'      
  )
  parser.add_argument(
      '--output_path',
      default=None,
      help='Path to write tracking results.',
  )
  parser.add_argument(
      '--guide_match_search_seconds',
      default=None,
      help='number of seconds to search the video for the best match with the guide template.',
  )
  parser.add_argument(
      '--microns_per_pixel',
      default=None,
      help='conversion factor for pixel distances to microns',
  )
  raw_args = parser.parse_args()

  if raw_args.json_config_path is not None:
    config = config_from_json(raw_args.json_config_path)
  else:
    config = config_from_cmdline(raw_args)

  runTrackTemplate(config)
